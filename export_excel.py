#-*- encoding: utf-8 -*-  
   
from openpyxl.workbook import Workbook     
from openpyxl.writer.excel import ExcelWriter   
from openpyxl.cell import get_column_letter
from openpyxl.reader.excel import load_workbook

# write results to excel file            
def writeExcel(targetFile,sheetName,cell,value):
    #读取excel2007文件 
    wb = load_workbook(targetFile)
    #新建一个excelWriter   
    ew = ExcelWriter(wb)
    
    #根据sheet的名称得到sheet
    ws = wb.get_sheet_by_name(sheetName)  

    #修改sheet的内容，可以按单元格指定，也可以根据数字坐标，使用get_column_letter方法得到横坐标的字母
    ws.cell(cell).value = value
    #保存文件
    ew.save(targetFile)
     
if  __name__ =="__main__":
    writeExcel("gttglobal.xlsx","Sheet","E2","debug")

